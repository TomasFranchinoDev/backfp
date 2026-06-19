"""
Generador dinámico de la plantilla Excel para importación SIU.

Crea un archivo .xlsx con 3 hojas:
  1. Instrucciones — guía paso a paso para la secretaria.
  2. Docentes     — datos del personal docente.
  3. Malla Académica — materias, carreras, horarios y asignaciones.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Estilos reutilizables ────────────────────────────────────────────────────

_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_HEADER_BORDER = Border(
    bottom=Side(style="thin", color="000000"),
    right=Side(style="thin", color="D9D9D9"),
)

_EXAMPLE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_EXAMPLE_FONT = Font(name="Calibri", italic=True, size=10, color="7F6000")

_TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
_TITLE_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="1F4E79")
_BODY_FONT = Font(name="Calibri", size=10, color="333333")
_TIP_FONT = Font(name="Calibri", size=10, color="2E75B6", italic=True)
_SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")


def _apply_header_row(ws, headers: list[str], widths: list[int]):
    """Aplica formato de encabezado y anchos de columna."""
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _HEADER_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _add_example_row(ws, row_num: int, values: list):
    """Agrega una fila de ejemplo con formato especial."""
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.fill = _EXAMPLE_FILL
        cell.font = _EXAMPLE_FONT


# ── Generación de hojas ─────────────────────────────────────────────────────

def _crear_hoja_instrucciones(wb: Workbook):
    ws = wb.active
    ws.title = "Instrucciones"
    ws.sheet_properties.tabColor = "1F4E79"

    # Configurar anchos
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 80

    # Título
    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value = "PLANTILLA DE IMPORTACIÓN SIU"
    title_cell.font = _TITLE_FONT
    title_cell.fill = _TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    instructions = [
        ("", ""),
        ("", "INSTRUCCIONES GENERALES"),
        ("1.", "Complete la hoja \"Docentes\" con los datos de todos los docentes a importar."),
        ("2.", "Complete la hoja \"Malla Académica\" con las materias, carreras, horarios y asignaciones."),
        ("3.", "Guarde el archivo y súbalo desde la sección de Importación del sistema."),
        ("4.", "El sistema validará TODOS los datos antes de guardar cualquier cosa."),
        ("5.", "Si hay errores, no se guardará nada. Corrija los errores y vuelva a subir."),
        ("", ""),
        ("", "HOJA: DOCENTES"),
        ("", "DNI → Número de documento del docente (se usa como nombre de usuario)."),
        ("", "Nombre → Nombre de pila."),
        ("", "Apellido → Apellido del docente."),
        ("", "Email → Correo electrónico. Si lo deja vacío, se genera automáticamente."),
        ("", "Contraseña → Contraseña de acceso. Si la deja vacía, se usa el DNI."),
        ("", ""),
        ("", "HOJA: MALLA ACADÉMICA"),
        ("", "Cada fila representa una combinación de materia + carrera + horario + docente."),
        ("", "Si una materia tiene varios horarios, repita la fila cambiando día/hora."),
        ("", "Si una materia tiene varios docentes, repita la fila cambiando DNI/Rol."),
        ("", ""),
        ("", "Campos de Materia:"),
        ("", "  Código Materia SIU → Código único de la materia en el sistema SIU."),
        ("", "  Nombre Materia → Nombre completo de la materia."),
        ("", "  Año Cursado → Año en que se cursa la materia (ej: 1, 2, 3)."),
        ("", ""),
        ("", "Campos de Carrera:"),
        ("", "  Institución → Nombre de la institución (ej: ICES, UCSE)."),
        ("", "  Código Carrera → Código identificador de la carrera."),
        ("", "  Nombre Carrera → Nombre completo de la carrera."),
        ("", "  Duración Carrera (años) → Cantidad de años que dura la carrera."),
        ("", "  Año Plan de Estudio → Año de la materia dentro del plan de estudios."),
        ("", ""),
        ("", "Campos de Horario (opcionales como grupo):"),
        ("", "  Día de Clase → Lunes, Martes, Miércoles, Jueves, Viernes o Sábado."),
        ("", "  Hora Inicio → Hora de inicio en formato HH:MM (ej: 08:00)."),
        ("", "  Hora Fin → Hora de fin en formato HH:MM (ej: 10:00)."),
        ("", ""),
        ("", "Campos de Asignación (opcionales — dejar vacío si no hay docente asignado):"),
        ("", "  DNI Docente → DNI del docente (debe existir en la hoja Docentes o en el sistema)."),
        ("", "  Rol Docente → Titular o Adjunto."),
        ("", "  Fecha Inicio Asignación → Formato DD/MM/AAAA."),
        ("", "  Fecha Fin Asignación → Formato DD/MM/AAAA (opcional)."),
        ("", ""),
        ("⚠️", "IMPORTANTE: No modifique los nombres de las hojas ni de las columnas."),
        ("⚠️", "IMPORTANTE: Las filas de ejemplo (amarillas) pueden ser borradas o reemplazadas."),
        ("⚠️", "IMPORTANTE: Los datos repetidos (misma materia en varias filas) deben ser consistentes."),
    ]

    for row_idx, (col_a, col_b) in enumerate(instructions, start=2):
        cell_a = ws.cell(row=row_idx, column=1, value=col_a)
        cell_b = ws.cell(row=row_idx, column=2, value=col_b)

        if col_b and col_b.startswith("HOJA:") or col_b.startswith("INSTRUCCIONES"):
            cell_b.font = _SUBTITLE_FONT
            cell_a.fill = _SECTION_FILL
            cell_b.fill = _SECTION_FILL
            ws.row_dimensions[row_idx].height = 24
        elif col_a.startswith("⚠️"):
            cell_b.font = _TIP_FONT
            cell_a.font = Font(size=12)
        else:
            cell_b.font = _BODY_FONT

    ws.sheet_view.showGridLines = False


def _crear_hoja_docentes(wb: Workbook):
    ws = wb.create_sheet("Docentes")
    ws.sheet_properties.tabColor = "2E75B6"

    headers = ["DNI", "Nombre", "Apellido", "Email", "Contraseña"]
    widths = [18, 22, 22, 32, 22]
    _apply_header_row(ws, headers, widths)

    # Filas de ejemplo
    _add_example_row(ws, 2, ["40111222", "Juan", "Pérez", "juan@ices.edu.ar", ""])
    _add_example_row(ws, 3, ["40333444", "María", "Gómez", "maria@ices.edu.ar", "MiClave123"])


def _crear_hoja_malla(wb: Workbook):
    ws = wb.create_sheet("Malla Académica")
    ws.sheet_properties.tabColor = "548235"

    headers = [
        "Código Materia SIU", "Nombre Materia", "Año Cursado",
        "Institución", "Código Carrera", "Nombre Carrera", "Duración Carrera (años)",
        "Año Plan de Estudio",
        "Día de Clase", "Hora Inicio", "Hora Fin",
        "DNI Docente", "Rol Docente",
        "Fecha Inicio Asignación", "Fecha Fin Asignación",
    ]
    widths = [22, 28, 14, 16, 18, 34, 22, 20, 16, 14, 14, 18, 16, 24, 24]
    _apply_header_row(ws, headers, widths)

    # Filas de ejemplo
    _add_example_row(ws, 2, [
        "MAT01", "Matemática I", 1,
        "ICES", "TPI", "Tecnicatura en Programación", 3, 1,
        "Lunes", "08:00", "10:00",
        "40111222", "Titular", "01/03/2026", "31/12/2026",
    ])
    _add_example_row(ws, 3, [
        "MAT01", "Matemática I", 1,
        "ICES", "TPI", "Tecnicatura en Programación", 3, 1,
        "Miércoles", "08:00", "10:00",
        "40111222", "Titular", "01/03/2026", "31/12/2026",
    ])
    _add_example_row(ws, 4, [
        "PROG1", "Programación I", 1,
        "ICES", "TPI", "Tecnicatura en Programación", 3, 1,
        "Martes", "14:30", "17:30",
        "40333444", "Adjunto", "01/03/2026", "",
    ])
    _add_example_row(ws, 5, [
        "PROG1", "Programación I", 1,
        "ICES", "TUI", "Tecnicatura en Informática", 2, 1,
        "Jueves", "10:00", "12:00",
        "", "", "", "",
    ])

    # ── Data Validation (dropdowns) ──────────────────────────────────────

    # Día de Clase
    dv_dia = DataValidation(
        type="list",
        formula1='"Lunes,Martes,Miércoles,Jueves,Viernes,Sábado"',
        allow_blank=True,
    )
    dv_dia.error = "Seleccione un día válido"
    dv_dia.errorTitle = "Día inválido"
    dv_dia.prompt = "Seleccione el día de clase"
    dv_dia.promptTitle = "Día de Clase"
    ws.add_data_validation(dv_dia)
    dv_dia.add(f"I2:I1000")

    # Rol Docente
    dv_rol = DataValidation(
        type="list",
        formula1='"Titular,Adjunto"',
        allow_blank=True,
    )
    dv_rol.error = "Seleccione Titular o Adjunto"
    dv_rol.errorTitle = "Rol inválido"
    dv_rol.prompt = "Seleccione el rol del docente"
    dv_rol.promptTitle = "Rol Docente"
    ws.add_data_validation(dv_rol)
    dv_rol.add(f"M2:M1000")

    ws.freeze_panes = "A2"


# ── API pública ──────────────────────────────────────────────────────────────

def generar_plantilla_excel() -> io.BytesIO:
    """Genera el archivo Excel plantilla y lo devuelve como BytesIO."""
    wb = Workbook()

    _crear_hoja_instrucciones(wb)
    _crear_hoja_docentes(wb)
    _crear_hoja_malla(wb)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
