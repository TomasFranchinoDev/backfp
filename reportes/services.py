import calendar
from datetime import date, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.utils import timezone

from calendario.models import EventoCalendario
from asignaciones.models import AsignacionDocente
from academico.models import SlotHorario
from asistencia.models import RegistroAsistencia
from core.constants import DiaSemana, TipoClase


def format_tipo_clase(tipo: str) -> str:
    """Retorna el nombre legible y normalizado del tipo de clase."""
    if not tipo:
        return "Presencial"
    tipo_clean = str(tipo).lower().strip()
    if tipo_clean in (TipoClase.VIRTUAL_SINCRONICA, 'virtual_sincronica', 'virtual sincrónica', 'virtual sincronica'):
        return "Virtual Sincrónica"
    elif tipo_clean in (TipoClase.ASINCRONICA, 'asincronica', 'asincrónica'):
        return "Asincrónica"
    return "Presencial"


def calcular_ausencias_dinamicas(desde: date, hasta: date, institucion: str = None, agrupar_por: str = 'docente'):
    """
    Cruza el catálogo teórico vs los registros reales para deducir asistencias y ausencias.
    Discrimina las asistencias por Tipo de Clase (Presencial, Virtual Sincrónica, Asincrónica).
    Soporta agrupamiento por 'docente', 'carrera' o 'materia'.
    """
    fecha_inicio = desde
    fecha_fin = hasta

    # 1. Obtener eventos del calendario académico del mes
    eventos = EventoCalendario.objects.filter(fecha__range=[fecha_inicio, fecha_fin])
    mapa_eventos = {evento.fecha: evento.descripcion for evento in eventos}

    # 2. Precargar todos los registros de asistencia del mes
    registros = RegistroAsistencia.objects.filter(fecha__range=[fecha_inicio, fecha_fin])
    # Clave: (docente_id, slot_id, fecha) -> Valor: Objeto RegistroAsistencia
    mapa_asistencia = {(r.docente_id, r.slot_horario_id, r.fecha): r for r in registros}

    # 3. Filtrar asignaciones activas (opcional por institución)
    asignaciones = AsignacionDocente.objects.filter(
        activa=True,
        fecha_inicio__lte=fecha_fin
    ).select_related('docente__user', 'materia')

    if institucion:
        asignaciones = asignaciones.filter(materia__carreras_asociadas__carrera__institucion=institucion).distinct()

    # Prefetch de carreras asociadas para evitar consultas N+1
    asignaciones = asignaciones.prefetch_related('materia__carreras_asociadas__carrera')

    # 4. Precargar Slots Horarios para esas materias
    materias_ids = [a.materia_id for a in asignaciones]
    slots = SlotHorario.objects.filter(materia_id__in=materias_ids).select_related('materia')
    
    # Agrupar slots por materia: {materia_id: [slot1, slot2]}
    slots_por_materia = defaultdict(list)
    for slot in slots:
        slots_por_materia[slot.materia_id].append(slot)

    # Estructura temporal para acumular datos
    reporte_grupos = defaultdict(lambda: {
        'nombre': '',
        'codigo': None,
        'esperadas': 0,
        'asistencias': 0,
        'asistencias_presenciales': 0,
        'asistencias_virtuales_sincronicas': 0,
        'asistencias_asincronicas': 0,
        'ausencias': [],
        'detalle_asistencias': [],
    })

    nombres_dias = dict(DiaSemana.choices)

    # Iterar sobre las fechas del rango
    delta_days = (fecha_fin - fecha_inicio).days + 1
    for dia in range(delta_days):
        fecha_actual = fecha_inicio + timedelta(days=dia)
        tiene_evento = fecha_actual in mapa_eventos
        evento_desc = mapa_eventos.get(fecha_actual)
        dia_semana_actual = fecha_actual.weekday() # 0 = Lunes

        for asignacion in asignaciones:
            # Check si la asignación cubría esta fecha específica
            if asignacion.fecha_inicio > fecha_actual:
                continue
            if asignacion.fecha_fin and asignacion.fecha_fin < fecha_actual:
                continue

            docente_id = asignacion.docente_id
            docente_nombre = asignacion.docente.user.get_full_name()
            materia_id = asignacion.materia_id
            materia_nombre = asignacion.materia.nombre
            materia_codigo = asignacion.materia.codigo_siu
            carreras_asoc = list(asignacion.materia.carreras_asociadas.all())

            # Revisar si hay clases programadas para este día de la semana, vigentes en la fecha actual
            slots_hoy = [
                s for s in slots_por_materia[materia_id] 
                if s.dia_semana == dia_semana_actual and s.is_valid_at(fecha_actual)
            ]

            for slot in slots_hoy:
                registro = mapa_asistencia.get((docente_id, slot.id, fecha_actual))
                asistio = registro is not None
                
                # Definir los grupos a los que afecta esta clase
                grupos = []
                if agrupar_por == 'docente':
                    grupos.append({
                        'id': docente_id,
                        'nombre': docente_nombre,
                        'codigo': None
                    })
                elif agrupar_por == 'carrera':
                    for mc in carreras_asoc:
                        grupos.append({
                            'id': mc.carrera.id,
                            'nombre': mc.carrera.nombre,
                            'codigo': mc.carrera.codigo
                        })
                elif agrupar_por == 'materia':
                    grupos.append({
                        'id': materia_id,
                        'nombre': materia_nombre,
                        'codigo': materia_codigo
                    })

                for grp in grupos:
                    gid = grp['id']
                    reporte_grupos[gid]['nombre'] = grp['nombre']
                    reporte_grupos[gid]['codigo'] = grp['codigo']

                    if asistio:
                        tipo_legible = format_tipo_clase(registro.tipo_clase)
                        reporte_grupos[gid]['asistencias'] += 1
                        if tipo_legible == "Presencial":
                            reporte_grupos[gid]['asistencias_presenciales'] += 1
                        elif tipo_legible == "Virtual Sincrónica":
                            reporte_grupos[gid]['asistencias_virtuales_sincronicas'] += 1
                        elif tipo_legible == "Asincrónica":
                            reporte_grupos[gid]['asistencias_asincronicas'] += 1

                        # Si asistió en un feriado/bloqueado, lo sumamos a esperadas para evitar % > 100
                        reporte_grupos[gid]['esperadas'] += 1

                        hora_ent = timezone.localtime(registro.hora_entrada).strftime('%H:%M') if registro.hora_entrada else None
                        hora_sal = timezone.localtime(registro.hora_salida).strftime('%H:%M') if registro.hora_salida else None

                        reporte_grupos[gid]['detalle_asistencias'].append({
                            'fecha': fecha_actual,
                            'materia_nombre': materia_nombre,
                            'docente_nombre': docente_nombre,
                            'dia_semana': nombres_dias.get(dia_semana_actual, str(dia_semana_actual)),
                            'hora_inicio': slot.hora_inicio.strftime('%H:%M'),
                            'hora_fin': slot.hora_fin.strftime('%H:%M'),
                            'tipo_clase': tipo_legible,
                            'hora_entrada': hora_ent,
                            'hora_salida': hora_sal,
                        })
                    else:
                        # Ausencia
                        if tiene_evento:
                            # Feriado: se registra con el evento pero no cuenta para la estadística de esperadas/ausencias
                            reporte_grupos[gid]['ausencias'].append({
                                'fecha': fecha_actual,
                                'materia_nombre': materia_nombre,
                                'docente_nombre': docente_nombre,
                                'dia_semana': nombres_dias.get(dia_semana_actual, str(dia_semana_actual)),
                                'hora_inicio': slot.hora_inicio.strftime('%H:%M'),
                                'evento_calendario': evento_desc
                            })
                        else:
                            # Normal: suma a esperadas y ausencias
                            reporte_grupos[gid]['esperadas'] += 1
                            reporte_grupos[gid]['ausencias'].append({
                                'fecha': fecha_actual,
                                'materia_nombre': materia_nombre,
                                'docente_nombre': docente_nombre,
                                'dia_semana': nombres_dias.get(dia_semana_actual, str(dia_semana_actual)),
                                'hora_inicio': slot.hora_inicio.strftime('%H:%M'),
                                'evento_calendario': None
                            })

    return reporte_grupos


def generar_datos_desnormalizados(desde: date, hasta: date, institucion: str = None):
    """
    Genera una lista plana de diccionarios con todas las inasistencias desnormalizadas.
    Cada elemento representa una única ausencia con todos los datos cruzados (14 columnas).
    Incluye ausencias por feriado con la columna 'tipo_dia' indicando el evento.
    """
    fecha_inicio = desde
    fecha_fin = hasta

    # 1. Eventos del calendario
    eventos = EventoCalendario.objects.filter(fecha__range=[fecha_inicio, fecha_fin])
    mapa_eventos = {evento.fecha: evento.descripcion for evento in eventos}

    # 2. Registros de asistencia
    registros = RegistroAsistencia.objects.filter(fecha__range=[fecha_inicio, fecha_fin])
    mapa_asistencia = {(r.docente_id, r.slot_horario_id, r.fecha): True for r in registros}

    # 3. Asignaciones activas
    asignaciones = AsignacionDocente.objects.filter(
        activa=True,
        fecha_inicio__lte=fecha_fin
    ).select_related('docente__user', 'materia')

    if institucion:
        asignaciones = asignaciones.filter(materia__carreras_asociadas__carrera__institucion=institucion).distinct()

    asignaciones = asignaciones.prefetch_related('materia__carreras_asociadas__carrera')

    # 4. Slots horarios
    materias_ids = [a.materia_id for a in asignaciones]
    slots = SlotHorario.objects.filter(materia_id__in=materias_ids).select_related('materia')

    slots_por_materia = defaultdict(list)
    for slot in slots:
        slots_por_materia[slot.materia_id].append(slot)

    nombres_dias = dict(DiaSemana.choices)
    filas = []

    delta_days = (fecha_fin - fecha_inicio).days + 1
    for dia in range(delta_days):
        fecha_actual = fecha_inicio + timedelta(days=dia)
        tiene_evento = fecha_actual in mapa_eventos
        evento_desc = mapa_eventos.get(fecha_actual)
        dia_semana_actual = fecha_actual.weekday()

        for asignacion in asignaciones:
            if asignacion.fecha_inicio > fecha_actual:
                continue
            if asignacion.fecha_fin and asignacion.fecha_fin < fecha_actual:
                continue

            docente = asignacion.docente
            docente_nombre = docente.user.get_full_name()
            docente_dni = docente.user.username
            materia = asignacion.materia
            carreras_asoc = list(materia.carreras_asociadas.all())

            # Concatenar carreras asociadas
            if carreras_asoc:
                carreras_nombres = ' / '.join(mc.carrera.nombre for mc in carreras_asoc)
                carreras_codigos = ' / '.join(mc.carrera.codigo for mc in carreras_asoc)
                instituciones = ' / '.join(
                    mc.carrera.get_institucion_display() if hasattr(mc.carrera, 'get_institucion_display') 
                    else mc.carrera.institucion.upper() 
                    for mc in carreras_asoc
                )
            else:
                carreras_nombres = 'Sin carrera asignada'
                carreras_codigos = '-'
                instituciones = '-'

            slots_hoy = [
                s for s in slots_por_materia[materia.id]
                if s.dia_semana == dia_semana_actual and s.is_valid_at(fecha_actual)
            ]

            for slot in slots_hoy:
                asistio = (docente.id, slot.id, fecha_actual) in mapa_asistencia

                if not asistio:
                    filas.append({
                        'fecha': fecha_actual,
                        'dia': nombres_dias.get(dia_semana_actual, str(dia_semana_actual)),
                        'tipo_dia': evento_desc if tiene_evento else 'Laborable',
                        'docente': docente_nombre,
                        'dni': docente_dni,
                        'materia': materia.nombre,
                        'codigo_materia': materia.codigo_siu,
                        'anio_materia': materia.anio,
                        'carreras': carreras_nombres,
                        'codigo_carrera': carreras_codigos,
                        'institucion': instituciones,
                        'hora_inicio': slot.hora_inicio.strftime('%H:%M'),
                        'hora_fin': slot.hora_fin.strftime('%H:%M'),
                        'rol_docente': asignacion.get_rol_display() if hasattr(asignacion, 'get_rol_display') else asignacion.rol.capitalize(),
                    })

    # Ordenar por fecha, luego por docente
    filas.sort(key=lambda f: (f['fecha'], f['docente'], f['hora_inicio']))
    return filas


def generar_datos_desnormalizados_asistencias(desde: date, hasta: date, institucion: str = None):
    """
    Genera una lista plana de diccionarios con todas las asistencias registradas desnormalizadas.
    Incluye la columna 'tipo_clase' ('Presencial', 'Virtual Sincrónica', 'Asincrónica').
    """
    fecha_inicio = desde
    fecha_fin = hasta

    registros = RegistroAsistencia.objects.filter(
        fecha__range=[fecha_inicio, fecha_fin]
    ).select_related(
        'docente__user',
        'slot_horario__materia'
    ).prefetch_related(
        'slot_horario__materia__carreras_asociadas__carrera',
        'docente__asignaciones'
    )

    if institucion:
        registros = registros.filter(
            slot_horario__materia__carreras_asociadas__carrera__institucion=institucion
        ).distinct()

    nombres_dias = dict(DiaSemana.choices)
    filas = []

    for reg in registros:
        docente = reg.docente
        docente_nombre = docente.user.get_full_name()
        docente_dni = docente.user.username
        materia = reg.slot_horario.materia
        slot = reg.slot_horario
        carreras_asoc = list(materia.carreras_asociadas.all())

        if carreras_asoc:
            carreras_nombres = ' / '.join(mc.carrera.nombre for mc in carreras_asoc)
            carreras_codigos = ' / '.join(mc.carrera.codigo for mc in carreras_asoc)
            instituciones = ' / '.join(
                mc.carrera.get_institucion_display() if hasattr(mc.carrera, 'get_institucion_display')
                else mc.carrera.institucion.upper()
                for mc in carreras_asoc
            )
        else:
            carreras_nombres = 'Sin carrera asignada'
            carreras_codigos = '-'
            instituciones = '-'

        asignacion = docente.asignaciones.filter(materia=materia, activa=True).first()
        rol_docente = 'Titular'
        if asignacion:
            rol_docente = asignacion.get_rol_display() if hasattr(asignacion, 'get_rol_display') else str(asignacion.rol).capitalize()

        hora_ent_str = timezone.localtime(reg.hora_entrada).strftime('%H:%M') if reg.hora_entrada else '-'
        hora_sal_str = timezone.localtime(reg.hora_salida).strftime('%H:%M') if reg.hora_salida else '-'
        dia_semana_num = reg.fecha.weekday()

        filas.append({
            'fecha': reg.fecha,
            'dia': nombres_dias.get(dia_semana_num, str(dia_semana_num)),
            'tipo_clase': format_tipo_clase(reg.tipo_clase),
            'docente': docente_nombre,
            'dni': docente_dni,
            'materia': materia.nombre,
            'codigo_materia': materia.codigo_siu,
            'anio_materia': materia.anio,
            'carreras': carreras_nombres,
            'codigo_carrera': carreras_codigos,
            'institucion': instituciones,
            'hora_inicio': slot.hora_inicio.strftime('%H:%M'),
            'hora_fin': slot.hora_fin.strftime('%H:%M'),
            'hora_entrada': hora_ent_str,
            'hora_salida': hora_sal_str,
            'rol_docente': rol_docente,
        })

    filas.sort(key=lambda f: (f['fecha'], f['docente'], f['hora_inicio']))
    return filas


def generar_excel_ausencias(datos_desnormalizados: list, desde: date, hasta: date, institucion: str, datos_asistencias: list = None):
    """
    Genera un archivo Excel (.xlsx) completo y profesional con 4 hojas:
    1. "Asistencias": Detalle desnormalizado con 'Tipo de Clase' (Presencial, Virtual Sincrónica, Asincrónica).
    2. "Inasistencias": Detalle desnormalizado de ausencias y feriados.
    3. "Resumen Consolidado": Resumen agrupado por docente con métricas discriminadas por tipo de clase.
    4. "Info Reporte": Metadatos y parámetros del reporte generado.
    """
    if datos_asistencias is None:
        datos_asistencias = generar_datos_desnormalizados_asistencias(desde, hasta, institucion)

    wb = Workbook()

    # ── Estilos Comunes ──
    header_fill = PatternFill(start_color="2B3A67", end_color="2B3A67", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    date_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    # Estilos por Tipo de Clase
    presencial_fill = PatternFill(start_color="EBF4FF", end_color="EBF4FF", fill_type="solid")
    presencial_font = Font(color="1E3A8A", bold=True)
    virtual_fill = PatternFill(start_color="F5F3FF", end_color="F5F3FF", fill_type="solid")
    virtual_font = Font(color="5B21B6", bold=True)
    asincronica_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    asincronica_font = Font(color="92400E", bold=True)

    feriado_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    feriado_font = Font(color="856404", italic=True)

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 1: Asistencias
    # ══════════════════════════════════════════════════════════════════════════
    ws_asist = wb.active
    ws_asist.title = "Asistencias"

    headers_asist = [
        "Fecha",
        "Día",
        "Tipo de Clase",
        "Docente",
        "DNI",
        "Materia",
        "Código Materia (SIU)",
        "Año Materia",
        "Carrera(s)",
        "Código Carrera",
        "Institución",
        "Hora Inicio Slot",
        "Hora Fin Slot",
        "Hora Entrada",
        "Hora Salida",
        "Rol Docente",
    ]

    for col_num, header in enumerate(headers_asist, 1):
        cell = ws_asist.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for row_idx, fila in enumerate(datos_asistencias, start=2):
        fecha_cell = ws_asist.cell(row=row_idx, column=1, value=fila['fecha'].strftime('%d/%m/%Y'))
        fecha_cell.font = date_font
        fecha_cell.alignment = center_align

        ws_asist.cell(row=row_idx, column=2, value=fila['dia'])
        tipo_cell = ws_asist.cell(row=row_idx, column=3, value=fila['tipo_clase'])
        tipo_cell.alignment = center_align

        # Resaltado sutil por tipo de clase
        if fila['tipo_clase'] == 'Presencial':
            tipo_cell.fill = presencial_fill
            tipo_cell.font = presencial_font
        elif fila['tipo_clase'] == 'Virtual Sincrónica':
            tipo_cell.fill = virtual_fill
            tipo_cell.font = virtual_font
        elif fila['tipo_clase'] == 'Asincrónica':
            tipo_cell.fill = asincronica_fill
            tipo_cell.font = asincronica_font

        ws_asist.cell(row=row_idx, column=4, value=fila['docente'])
        ws_asist.cell(row=row_idx, column=5, value=fila['dni'])
        ws_asist.cell(row=row_idx, column=6, value=fila['materia'])
        ws_asist.cell(row=row_idx, column=7, value=fila['codigo_materia'])
        ws_asist.cell(row=row_idx, column=8, value=fila['anio_materia']).alignment = center_align
        ws_asist.cell(row=row_idx, column=9, value=fila['carreras'])
        ws_asist.cell(row=row_idx, column=10, value=fila['codigo_carrera'])
        ws_asist.cell(row=row_idx, column=11, value=fila['institucion'])
        ws_asist.cell(row=row_idx, column=12, value=fila['hora_inicio']).alignment = center_align
        ws_asist.cell(row=row_idx, column=13, value=fila['hora_fin']).alignment = center_align
        ws_asist.cell(row=row_idx, column=14, value=fila['hora_entrada']).alignment = center_align
        ws_asist.cell(row=row_idx, column=15, value=fila['hora_salida']).alignment = center_align
        ws_asist.cell(row=row_idx, column=16, value=fila['rol_docente'])

    last_row_asist = max(len(datos_asistencias) + 1, 2)
    ws_asist.auto_filter.ref = f"A1:P{last_row_asist}"
    ws_asist.freeze_panes = "A2"

    column_widths_asist = {
        'A': 14,   # Fecha
        'B': 12,   # Día
        'C': 20,   # Tipo de Clase
        'D': 30,   # Docente
        'E': 14,   # DNI
        'F': 35,   # Materia
        'G': 20,   # Código Materia
        'H': 14,   # Año Materia
        'I': 35,   # Carrera(s)
        'J': 16,   # Código Carrera
        'K': 14,   # Institución
        'L': 16,   # Hora Inicio Slot
        'M': 16,   # Hora Fin Slot
        'N': 14,   # Hora Entrada
        'O': 14,   # Hora Salida
        'P': 14,   # Rol Docente
    }
    for col_letter, width in column_widths_asist.items():
        ws_asist.column_dimensions[col_letter].width = width

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 2: Inasistencias
    # ══════════════════════════════════════════════════════════════════════════
    ws_inasist = wb.create_sheet(title="Inasistencias")

    headers_inasist = [
        "Fecha",
        "Día",
        "Tipo Día",
        "Docente",
        "DNI",
        "Materia",
        "Código Materia (SIU)",
        "Año Materia",
        "Carrera(s)",
        "Código Carrera",
        "Institución",
        "Hora Inicio",
        "Hora Fin",
        "Rol Docente",
    ]

    for col_num, header in enumerate(headers_inasist, 1):
        cell = ws_inasist.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for row_idx, fila in enumerate(datos_desnormalizados, start=2):
        fecha_cell = ws_inasist.cell(row=row_idx, column=1, value=fila['fecha'].strftime('%d/%m/%Y'))
        fecha_cell.font = date_font
        fecha_cell.alignment = center_align

        ws_inasist.cell(row=row_idx, column=2, value=fila['dia'])
        tipo_dia_cell = ws_inasist.cell(row=row_idx, column=3, value=fila['tipo_dia'])
        ws_inasist.cell(row=row_idx, column=4, value=fila['docente'])
        ws_inasist.cell(row=row_idx, column=5, value=fila['dni'])
        ws_inasist.cell(row=row_idx, column=6, value=fila['materia'])
        ws_inasist.cell(row=row_idx, column=7, value=fila['codigo_materia'])
        ws_inasist.cell(row=row_idx, column=8, value=fila['anio_materia']).alignment = center_align
        ws_inasist.cell(row=row_idx, column=9, value=fila['carreras'])
        ws_inasist.cell(row=row_idx, column=10, value=fila['codigo_carrera'])
        ws_inasist.cell(row=row_idx, column=11, value=fila['institucion'])
        ws_inasist.cell(row=row_idx, column=12, value=fila['hora_inicio']).alignment = center_align
        ws_inasist.cell(row=row_idx, column=13, value=fila['hora_fin']).alignment = center_align
        ws_inasist.cell(row=row_idx, column=14, value=fila['rol_docente'])

        if fila['tipo_dia'] != 'Laborable':
            tipo_dia_cell.fill = feriado_fill
            tipo_dia_cell.font = feriado_font

    last_row_inasist = max(len(datos_desnormalizados) + 1, 2)
    ws_inasist.auto_filter.ref = f"A1:N{last_row_inasist}"
    ws_inasist.freeze_panes = "A2"

    column_widths_inasist = {
        'A': 14,   # Fecha
        'B': 12,   # Día
        'C': 28,   # Tipo Día
        'D': 30,   # Docente
        'E': 14,   # DNI
        'F': 35,   # Materia
        'G': 20,   # Código Materia
        'H': 14,   # Año Materia
        'I': 35,   # Carrera(s)
        'J': 16,   # Código Carrera
        'K': 14,   # Institución
        'L': 14,   # Hora Inicio
        'M': 12,   # Hora Fin
        'N': 14,   # Rol Docente
    }
    for col_letter, width in column_widths_inasist.items():
        ws_inasist.column_dimensions[col_letter].width = width

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 3: Resumen Consolidado
    # ══════════════════════════════════════════════════════════════════════════
    ws_resumen = wb.create_sheet(title="Resumen Consolidado")
    
    headers_resumen = [
        "Docente / Entidad",
        "Clases Esperadas",
        "Total Asistencias",
        "Presenciales",
        "Virtuales Sincrónicas",
        "Asincrónicas",
        "Ausencias",
        "% Presencia",
    ]

    for col_num, header in enumerate(headers_resumen, 1):
        cell = ws_resumen.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Calcular resumen por docente
    resumen_docentes = calcular_ausencias_dinamicas(desde, hasta, institucion, agrupar_por="docente")
    
    row_curr = 2
    for gid, d in sorted(resumen_docentes.items(), key=lambda x: x[1]['nombre']):
        ausencias_reales = len([a for a in d['ausencias'] if not a.get('evento_calendario')])
        pct = (d['asistencias'] / d['esperadas'] * 100) if d['esperadas'] > 0 else 0

        ws_resumen.cell(row=row_curr, column=1, value=d['nombre']).font = Font(bold=True)
        ws_resumen.cell(row=row_curr, column=2, value=d['esperadas']).alignment = center_align
        ws_resumen.cell(row=row_curr, column=3, value=d['asistencias']).alignment = center_align
        ws_resumen.cell(row=row_curr, column=4, value=d['asistencias_presenciales']).alignment = center_align
        ws_resumen.cell(row=row_curr, column=5, value=d['asistencias_virtuales_sincronicas']).alignment = center_align
        ws_resumen.cell(row=row_curr, column=6, value=d['asistencias_asincronicas']).alignment = center_align
        ws_resumen.cell(row=row_curr, column=7, value=ausencias_reales).alignment = center_align
        
        pct_cell = ws_resumen.cell(row=row_curr, column=8, value=f"{pct:.1f}%")
        pct_cell.alignment = center_align
        if pct >= 80:
            pct_cell.font = Font(color="166534", bold=True)
        elif pct >= 60:
            pct_cell.font = Font(color="854D0E", bold=True)
        else:
            pct_cell.font = Font(color="991B1B", bold=True)

        row_curr += 1

    last_row_resumen = max(row_curr - 1, 2)
    ws_resumen.auto_filter.ref = f"A1:H{last_row_resumen}"
    ws_resumen.freeze_panes = "A2"

    column_widths_resumen = {
        'A': 32,   # Docente
        'B': 18,   # Esperadas
        'C': 18,   # Total Asistencias
        'D': 16,   # Presenciales
        'E': 22,   # Virtuales Sincrónicas
        'F': 16,   # Asincrónicas
        'G': 14,   # Ausencias
        'H': 16,   # % Presencia
    }
    for col_letter, width in column_widths_resumen.items():
        ws_resumen.column_dimensions[col_letter].width = width

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 4: Info Reporte
    # ══════════════════════════════════════════════════════════════════════════
    ws_info = wb.create_sheet(title="Info Reporte")
    
    total_presenciales = sum(1 for a in datos_asistencias if a['tipo_clase'] == 'Presencial')
    total_virtuales = sum(1 for a in datos_asistencias if a['tipo_clase'] == 'Virtual Sincrónica')
    total_asincronicas = sum(1 for a in datos_asistencias if a['tipo_clase'] == 'Asincrónica')
    total_inasist_laborables = sum(1 for f in datos_desnormalizados if f['tipo_dia'] == 'Laborable')
    total_inasist_feriado = sum(1 for f in datos_desnormalizados if f['tipo_dia'] != 'Laborable')

    info_data = [
        ("Parámetro", "Valor"),
        ("Fecha Desde", desde.strftime('%d/%m/%Y')),
        ("Fecha Hasta", hasta.strftime('%d/%m/%Y')),
        ("Institución", institucion if institucion else "Todas"),
        ("Total Clases Dictadas / Asistencias", len(datos_asistencias)),
        ("  - Presenciales", total_presenciales),
        ("  - Virtuales Sincrónicas", total_virtuales),
        ("  - Asincrónicas", total_asincronicas),
        ("Total Inasistencias", len(datos_desnormalizados)),
        ("  - Inasistencias Laborables", total_inasist_laborables),
        ("  - Inasistencias en Feriado/Evento", total_inasist_feriado),
    ]

    for row_idx, (param, valor) in enumerate(info_data, start=1):
        cell_param = ws_info.cell(row=row_idx, column=1, value=param)
        cell_valor = ws_info.cell(row=row_idx, column=2, value=valor)
        if row_idx == 1:
            cell_param.fill = header_fill
            cell_param.font = header_font
            cell_valor.fill = header_fill
            cell_valor.font = header_font
        else:
            if "  -" in str(param):
                cell_param.font = Font(italic=True)
            else:
                cell_param.font = Font(bold=True)

    ws_info.column_dimensions['A'].width = 38
    ws_info.column_dimensions['B'].width = 25

    return wb